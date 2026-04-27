"""
face_count.py
顔認識でファイルごとに人物をカウントしExcel出力

バージョン: 1.0.0 (2026-04-27)
作者: Yossy / Claude
"""

# ========================================
# パラメータ設定
# ========================================
TARGET_FOLDER        = r"\\IESSANsNAS\nas_share\Photo\2025\2503"
FACE_DB_FOLDER       = r"C:\GitHub\test\face_filter\face_database"
OUTPUT_FOLDER        = r"C:\GitHub\test\face_filter\output"
OUTPUT_DEBUG_FOLDER  = r"C:\GitHub\test\face_filter\output_debug"

# デバッグ用：顔のbbox+名前を描画した画像を OUTPUT_DEBUG_FOLDER に保存する
DEBUG_DRAW_ANNOTATIONS = True
DEBUG_FONT_PATH        = r"C:\Windows\Fonts\meiryo.ttc"

# 同一人物と判定する類似度しきい値（コサイン類似度。0〜1。高いほど厳しい）
# 推奨: 0.35〜0.50。集合写真で取りこぼしが多ければ下げる、誤検出が多ければ上げる
SIMILARITY_THRESHOLD = 0.60

# 検出時の画像最大サイズ（長辺ピクセル。小さいほど高速だが小さい顔を逃す）
DETECTION_MAX_SIZE   = 1280

# 検出器の入力サイズ（大きいほど小さい顔を捉えるが遅い）
DETECTION_INPUT_SIZE = (640, 640)

# DB登録時のフォールバック検出しきい値（先頭から順に試す。赤ちゃん・横顔の救済用）
DB_DETECTION_FALLBACK_THRESHOLDS = [0.5, 0.3, 0.1]

# サブフォルダも対象にするか
RECURSIVE            = False

# 処理する最大ファイル数（None=全件、整数=先頭からその件数のみ。動作確認・テスト用）
MAX_FILES            = 20

# 対象拡張子（小文字）
EXTENSIONS           = ['.jpg', '.jpeg', '.png', '.heic', '.bmp', '.webp']

# InsightFaceモデル: 'buffalo_l'(高精度・約300MB) / 'buffalo_s'(軽量・約16MB)
MODEL_NAME           = 'buffalo_l'

# ========================================
# imports
# ========================================
import sys
from pathlib import Path
from datetime import datetime
import numpy as np
import cv2
from PIL import Image, ImageOps, ImageDraw, ImageFont
import pillow_heif
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

pillow_heif.register_heif_opener()


# ========================================
# 認識エンジン抽象化（後で別ライブラリに差し替え可能にするため）
# ========================================
class FaceRecognizer:
    def get_embeddings(self, img):
        raise NotImplementedError

    def similarity(self, a, b):
        raise NotImplementedError


class InsightFaceRecognizer(FaceRecognizer):
    def __init__(self, model_name, det_size):
        from insightface.app import FaceAnalysis
        self.model_name = model_name
        self.det_size = det_size
        self.app = FaceAnalysis(name=model_name, providers=['CPUExecutionProvider'])
        self.app.prepare(ctx_id=0, det_size=det_size)

    def get_embeddings(self, img):
        faces = self.app.get(img)
        return [(face.normed_embedding, face.bbox) for face in faces]

    def get_embeddings_with_fallback(self, img, thresholds):
        # DB構築時の検出失敗を救済するため検出しきい値を段階的に下げて再試行する
        from insightface.app import FaceAnalysis
        for thresh in thresholds:
            app = FaceAnalysis(name=self.model_name, providers=['CPUExecutionProvider'])
            app.prepare(ctx_id=0, det_thresh=thresh, det_size=self.det_size)
            faces = app.get(img)
            if faces:
                return [(f.normed_embedding, f.bbox) for f in faces], thresh
        return [], None

    def similarity(self, a, b):
        # normed_embeddingはL2正規化済みのため内積=コサイン類似度
        return float(np.dot(a, b))


# ========================================
# 画像読み込み（HEIC・日本語パス対応）
# ========================================
def load_image(path):
    # スマホ写真のEXIF Orientationを反映するためPIL経由で読む（cv2はEXIFを無視するので横倒し画像になる）
    p = Path(path)
    img_pil = Image.open(p)
    img_pil = ImageOps.exif_transpose(img_pil)
    if img_pil.mode == 'RGBA':
        # 透過PNGは白背景に合成（透明部分を0埋めすると検出器が誤動作しやすい）
        bg = Image.new('RGB', img_pil.size, (255, 255, 255))
        bg.paste(img_pil, mask=img_pil.split()[3])
        img_pil = bg
    elif img_pil.mode != 'RGB':
        img_pil = img_pil.convert('RGB')
    return cv2.cvtColor(np.array(img_pil), cv2.COLOR_RGB2BGR)


def resize_for_detection(img, max_size):
    h, w = img.shape[:2]
    if max(h, w) <= max_size:
        return img
    scale = max_size / max(h, w)
    return cv2.resize(img, (int(w * scale), int(h * scale)))


# ========================================
# 顔データベース構築
# ========================================
def label_from_filename(stem):
    # 「クラス名_名前(_補助)?」の最初の2セグメントを人物ラベルとする。残りは同一人物の追加サンプル扱い
    parts = stem.split('_', 2)
    if len(parts) >= 2:
        return f"{parts[0]}_{parts[1]}"
    return stem


def build_face_database(recognizer, db_folder, extensions, fallback_thresholds):
    # ラベル単位で複数枚の埋め込みをまとめる
    groups = {}
    for img_path in sorted(Path(db_folder).iterdir()):
        if img_path.suffix.lower() not in extensions:
            continue
        label = label_from_filename(img_path.stem)
        img = load_image(img_path)
        embeddings, used_thresh = recognizer.get_embeddings_with_fallback(img, fallback_thresholds)
        if not embeddings:
            print(f"  [WARN] 顔検出失敗（しきい値{fallback_thresholds[-1]}でも検出不可）: {img_path.name}")
            continue
        if len(embeddings) > 1:
            embeddings.sort(
                key=lambda e: (e[1][2] - e[1][0]) * (e[1][3] - e[1][1]),
                reverse=True
            )
            print(f"  [WARN] 複数顔検出（最大の顔を使用）: {img_path.name}")
        groups.setdefault(label, []).append(embeddings[0][0])
        note = f"（しきい値={used_thresh}）" if used_thresh != fallback_thresholds[0] else ""
        print(f"  読込: {img_path.name} → {label} {note}")

    # 同一ラベル複数枚は埋め込みを平均化して再正規化（コサイン類似度のため再L2正規化が必要）
    db = {}
    for label, embs in groups.items():
        if len(embs) == 1:
            db[label] = embs[0]
        else:
            avg = np.mean(embs, axis=0)
            avg = avg / np.linalg.norm(avg)
            db[label] = avg
            print(f"  集約: {label} ← {len(embs)}枚平均")
    return db


# ========================================
# 対象ファイル収集
# ========================================
def collect_target_files(folder, recursive, extensions):
    p = Path(folder)
    if recursive:
        files = [f for f in p.rglob('*') if f.is_file() and f.suffix.lower() in extensions]
    else:
        files = [f for f in p.iterdir() if f.is_file() and f.suffix.lower() in extensions]
    return sorted(files)


# ========================================
# 認識処理（1画像分）
# ========================================
def recognize(recognizer, img, db, threshold):
    embeddings = recognizer.get_embeddings(img)
    matched = set()
    face_results = []  # [(bbox, matched_label_or_None, best_sim), ...]
    for emb, bbox in embeddings:
        best_label, best_sim = None, -1.0
        for label, db_emb in db.items():
            sim = recognizer.similarity(emb, db_emb)
            if sim > best_sim:
                best_sim = sim
                best_label = label
        if best_sim >= threshold:
            matched.add(best_label)
            face_results.append((bbox, best_label, best_sim))
        else:
            face_results.append((bbox, None, best_sim))
    return matched, face_results


# ========================================
# デバッグ用画像注釈
# ========================================
def draw_annotated_image(img_bgr, face_results, output_path, font_path):
    # 認識結果が正しいか目視確認するため、顔のbboxと判定名を描画して保存
    img_pil = Image.fromarray(cv2.cvtColor(img_bgr, cv2.COLOR_BGR2RGB))
    draw = ImageDraw.Draw(img_pil)

    for bbox, label, sim in face_results:
        x1, y1, x2, y2 = [int(v) for v in bbox]
        face_height = max(1, y2 - y1)
        # フォントサイズ・線幅は顔サイズに比例（画像全体ではなく顔基準のほうが視認性が安定する）
        font_size = max(10, min(28, face_height // 7))
        line_width = max(1, face_height // 100)
        font = ImageFont.truetype(font_path, font_size)

        is_matched = label is not None
        color = (0, 220, 0) if is_matched else (255, 60, 60)
        draw.rectangle([x1, y1, x2, y2], outline=color, width=line_width)

        text = f"{label} ({sim:.2f})" if is_matched else f"? ({sim:.2f})"
        try:
            bbox_text = draw.textbbox((0, 0), text, font=font)
            tw = bbox_text[2] - bbox_text[0]
            th = bbox_text[3] - bbox_text[1]
        except Exception:
            tw, th = font_size * len(text) // 2, font_size
        # テキストはbboxの上に置くが、画面外にはみ出す場合は下に
        ty = y1 - th - 4 if y1 - th - 4 >= 0 else y2 + 4
        draw.rectangle([x1, ty, x1 + tw + 4, ty + th + 4], fill=color)
        draw.text((x1 + 2, ty + 2), text, fill=(255, 255, 255), font=font)

    img_pil.save(output_path)


# ========================================
# Excel出力
# ========================================
def write_excel(output_path, results, labels):
    wb = Workbook()
    ws = wb.active
    ws.title = "face_count"

    bold = Font(bold=True)
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # ヘッダー
    c = ws.cell(row=1, column=1, value="ファイル名")
    c.font = bold; c.fill = header_fill
    for col, label in enumerate(labels, start=2):
        c = ws.cell(row=1, column=col, value=label)
        c.font = bold; c.fill = header_fill
        c.alignment = Alignment(textRotation=90, horizontal='center', vertical='bottom')
    total_col = len(labels) + 2
    c = ws.cell(row=1, column=total_col, value="合計人数")
    c.font = bold; c.fill = total_fill
    c.alignment = Alignment(textRotation=90, horizontal='center', vertical='bottom')

    # データ行
    filenames = sorted(results.keys())
    last_label_col_letter = get_column_letter(total_col - 1)
    for row_idx, fname in enumerate(filenames, start=2):
        ws.cell(row=row_idx, column=1, value=fname)
        matched = results[fname]
        for col, label in enumerate(labels, start=2):
            if label in matched:
                c = ws.cell(row=row_idx, column=col, value="○")
                c.alignment = Alignment(horizontal='center')
        ws.cell(
            row=row_idx, column=total_col,
            value=f'=COUNTIF(B{row_idx}:{last_label_col_letter}{row_idx},"○")'
        ).fill = total_fill

    # 集計行（最終行）
    total_row = len(filenames) + 2
    c = ws.cell(row=total_row, column=1, value="合計")
    c.font = bold; c.fill = total_fill
    for col in range(2, total_col):
        col_letter = get_column_letter(col)
        c = ws.cell(
            row=total_row, column=col,
            value=f'=COUNTIF({col_letter}2:{col_letter}{total_row - 1},"○")'
        )
        c.font = bold; c.fill = total_fill

    # 列幅・行高さ・フリーズ
    ws.column_dimensions['A'].width = 35
    for col in range(2, total_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 6
    ws.row_dimensions[1].height = 120
    ws.freeze_panes = 'B2'

    wb.save(output_path)


# ========================================
# メイン処理
# ========================================
def main():
    print("=" * 60)
    print("顔認識による人物カウント")
    print("=" * 60)

    if not Path(TARGET_FOLDER).exists():
        print(f"[ERROR] 対象フォルダが存在しない: {TARGET_FOLDER}")
        sys.exit(1)
    if not Path(FACE_DB_FOLDER).exists():
        print(f"[ERROR] 顔DBフォルダが存在しない: {FACE_DB_FOLDER}")
        sys.exit(1)
    Path(OUTPUT_FOLDER).mkdir(parents=True, exist_ok=True)
    if DEBUG_DRAW_ANNOTATIONS:
        Path(OUTPUT_DEBUG_FOLDER).mkdir(parents=True, exist_ok=True)

    print("\n[1/4] 認識エンジン初期化中...")
    recognizer = InsightFaceRecognizer(model_name=MODEL_NAME, det_size=DETECTION_INPUT_SIZE)

    print(f"\n[2/4] 顔データベース読み込み: {FACE_DB_FOLDER}")
    db = build_face_database(recognizer, FACE_DB_FOLDER, EXTENSIONS, DB_DETECTION_FALLBACK_THRESHOLDS)
    if not db:
        print("[ERROR] 顔DBが空")
        sys.exit(1)
    labels = sorted(db.keys())
    print(f"  → {len(labels)}名 登録完了")

    print(f"\n[3/4] 対象ファイル収集: {TARGET_FOLDER}")
    files = collect_target_files(TARGET_FOLDER, RECURSIVE, EXTENSIONS)
    print(f"  → {len(files)}件")
    if not files:
        print("[ERROR] 対象ファイルなし")
        sys.exit(1)
    if MAX_FILES is not None and MAX_FILES < len(files):
        files = files[:MAX_FILES]
        print(f"  → MAX_FILES={MAX_FILES} のため先頭{MAX_FILES}件のみ処理")

    print(f"\n[4/4] 顔認識処理中（しきい値={SIMILARITY_THRESHOLD}）...")
    results = {}
    failed = []
    for i, f in enumerate(files, 1):
        try:
            img = load_image(f)
            img = resize_for_detection(img, DETECTION_MAX_SIZE)
            matched, face_results = recognize(recognizer, img, db, SIMILARITY_THRESHOLD)
            results[f.name] = matched
            print(f"  [{i:>4}/{len(files)}] {f.name}: {len(face_results)}顔 → {len(matched)}名一致")
            if DEBUG_DRAW_ANNOTATIONS:
                debug_path = Path(OUTPUT_DEBUG_FOLDER) / f.name
                draw_annotated_image(img, face_results, debug_path, DEBUG_FONT_PATH)
        except Exception as e:
            print(f"  [{i:>4}/{len(files)}] {f.name}: ERROR {e}")
            failed.append(f.name)
            results[f.name] = set()

    output_filename = f"face_count_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = Path(OUTPUT_FOLDER) / output_filename
    write_excel(output_path, results, labels)

    print("\n" + "=" * 60)
    print(f"出力: {output_path}")
    print(f"成功: {len(files) - len(failed)} / {len(files)} 件 / 失敗: {len(failed)} 件")
    if failed:
        print("\n失敗したファイル:")
        for f in failed:
            print(f"  - {f}")


if __name__ == '__main__':
    main()
